"""
Tally Module — Research Survey & Test Tallying Tool
Matches the Cronbach / Pearson r module aesthetic (dark sidebar, teal accent).

Features:
─────────────────────────────────────────────────────────────────────────────
PART 1  — Likert Scale Tallying
  • Click buttons to tally each scale point per item
  • Undo last tally per item
  • Auto-computes: Frequency Total, Weighted Mean, Interpretation
  • Configurable scale (4-point / 5-point / 7-point)
  • Custom scale labels (SA, A, D, SD etc.)

PART 11 — Test Item Tallying (Correct / Wrong)
  • Click Correct or Wrong per item
  • Undo supported
  • Auto-computes: Total, % Correct, % Wrong, Difficulty Index

GLOBAL AUTO-COMPUTATIONS
  • Overall / Grand Mean (Likert)
  • Total Correct / Total Wrong (Test)
  • Live running totals update on every click

EXPORT
  • Export to Excel (.xlsx) — formatted, two sheets:
      Sheet 1: Likert Tally Table
      Sheet 2: Test Item Tally Table
  • Session Save / Load (JSON)
  • Reset per item or Reset All
─────────────────────────────────────────────────────────────────────────────
"""

import os, sys, json, copy
from datetime import datetime

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ─── Palette ──────────────────────────────────────────────────────────────────
from ui_theme import (
    BG_DEEP, BG_CARD, BG_PANEL, BG_INPUT,
    ACCENT, ACCENT2, DANGER, WARN, SUCCESS, PURPLE,
    TEXT_PRI, TEXT_SEC, BORDER,
)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

ORANGE   = "#f97316"

FONT_HEAD = ("Segoe UI", 22, "bold")
FONT_CARD = ("Segoe UI", 15, "bold")
FONT_BODY = ("Segoe UI", 13)
FONT_MONO = ("Consolas", 12)
FONT_BTN  = ("Segoe UI", 13, "bold")
FONT_TINY = ("Segoe UI", 11)
FONT_SML  = ("Segoe UI", 10)

# ─── Scale Configs ────────────────────────────────────────────────────────────
SCALE_CONFIGS = {
    "4-point": {
        "points": 4,
        "labels": ["4 - Strongly Agree", "3 - Agree",
                   "2 - Disagree", "1 - Strongly Disagree"],
        "values": [4, 3, 2, 1],
        "colors": [SUCCESS, ACCENT, WARN, DANGER],
        "interp": [(3.50, 4.00, "Strongly Agree"),
                   (2.50, 3.49, "Agree"),
                   (1.50, 2.49, "Disagree"),
                   (1.00, 1.49, "Strongly Disagree")],
    },
    "5-point": {
        "points": 5,
        "labels": ["5 - Strongly Agree", "4 - Agree", "3 - Neutral",
                   "2 - Disagree", "1 - Strongly Disagree"],
        "values": [5, 4, 3, 2, 1],
        "colors": [SUCCESS, ACCENT, ACCENT2, WARN, DANGER],
        "interp": [(4.50, 5.00, "Strongly Agree"),
                   (3.50, 4.49, "Agree"),
                   (2.50, 3.49, "Neutral"),
                   (1.50, 2.49, "Disagree"),
                   (1.00, 1.49, "Strongly Disagree")],
    },
    "7-point": {
        "points": 7,
        "labels": ["7-Strongly Agree","6-Agree","5-Somewhat Agree",
                   "4-Neutral","3-Somewhat Disagree","2-Disagree",
                   "1-Strongly Disagree"],
        "values": [7, 6, 5, 4, 3, 2, 1],
        "colors": [SUCCESS, ACCENT, "#34d399", ACCENT2, WARN, ORANGE, DANGER],
        "interp": [(6.00, 7.00, "Strongly Agree"),
                   (5.00, 5.99, "Agree"),
                   (4.00, 4.99, "Somewhat Agree"),
                   (3.00, 3.99, "Neutral"),
                   (2.00, 2.99, "Somewhat Disagree"),
                   (1.00, 1.99, "Disagree"),
                   (0.00, 0.99, "Strongly Disagree")],
    },
}

# ─── UI Helpers ───────────────────────────────────────────────────────────────

def divider(parent):
    ctk.CTkFrame(parent, height=1, fg_color=BORDER,
                 corner_radius=0).pack(fill="x")

def styled_entry(parent, placeholder="", width=0, height=34):
    kw = dict(placeholder_text=placeholder, fg_color=BG_INPUT,
              border_color=BORDER, text_color=TEXT_PRI,
              placeholder_text_color=TEXT_SEC, border_width=1,
              corner_radius=6, height=height, font=FONT_BODY)
    e = ctk.CTkEntry(parent, **kw)
    if width: e.configure(width=width)
    return e

def sidebar_btn(parent, text, fg, hover, text_color=TEXT_PRI,
                font=FONT_BTN, height=36, state="normal"):
    return ctk.CTkButton(parent, text=text, fg_color=fg, hover_color=hover,
                         text_color=text_color, font=font, height=height,
                         corner_radius=8, state=state)

def card(parent, title="", **kw):
    f = ctk.CTkFrame(parent, fg_color=BG_CARD, corner_radius=12,
                     border_width=1, border_color=BORDER, **kw)
    if title:
        ctk.CTkLabel(f, text=title, font=FONT_CARD,
                     text_color=TEXT_PRI).pack(anchor="w", padx=16, pady=(12,4))
    return f

def sec_label(parent, text):
    ctk.CTkLabel(parent, text=text, font=("Segoe UI",11,"bold"),
                 text_color=TEXT_SEC).pack(anchor="w", padx=18, pady=(12,3))


# ─── Interpretation Helper ────────────────────────────────────────────────────

def get_interpretation(wm, scale_key):
    cfg = SCALE_CONFIGS[scale_key]
    for lo, hi, label in cfg["interp"]:
        if lo <= round(wm, 2) <= hi:
            return label
    return "—"


# ─── Excel Exporter ───────────────────────────────────────────────────────────

class ExcelExporter:

    HEADER_FILL   = PatternFill("solid", fgColor="1e2736")
    SUBHEAD_FILL  = PatternFill("solid", fgColor="161b22")
    ACCENT_FILL   = PatternFill("solid", fgColor="004d3d")
    ALT_FILL      = PatternFill("solid", fgColor="0d1117")
    TOTAL_FILL    = PatternFill("solid", fgColor="1c2230")
    GRAND_FILL    = PatternFill("solid", fgColor="002d22")

    WHITE    = Font(color="E6EDF3", name="Arial", size=10)
    BOLD_W   = Font(color="E6EDF3", name="Arial", size=10, bold=True)
    ACCENT_F = Font(color="00c9a7", name="Arial", size=10, bold=True)
    WARN_F   = Font(color="f59e0b", name="Arial", size=10, bold=True)
    DANGER_F = Font(color="ef4444", name="Arial", size=10, bold=True)
    SUCCESS_F= Font(color="22c55e", name="Arial", size=10, bold=True)

    THIN = Side(style="thin", color="30363d")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    @classmethod
    def _cell(cls, ws, row, col, value="", font=None, fill=None,
              align=None, border=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or cls.WHITE
        c.fill      = fill  or cls.SUBHEAD_FILL
        c.alignment = align or cls.CENTER
        if border: c.border = cls.BORDER
        return c

    @classmethod
    def export(cls, filename, likert_data, test_data,
               likert_cfg, n_respondents_likert,
               n_respondents_test, title="Tally Report",
               researcher="", subtitle=""):
        wb = openpyxl.Workbook()

        # ── Sheet 1: Likert ───────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Part 1 — Likert"
        cls._build_likert_sheet(ws1, likert_data, likert_cfg,
                                n_respondents_likert, title, researcher, subtitle)

        # ── Sheet 2: Test Items ───────────────────────────────────────────────
        ws2 = wb.create_sheet("Part 11 — Test Items")
        cls._build_test_sheet(ws2, test_data, n_respondents_test,
                              title, researcher, subtitle)

        wb.save(filename)

    # ── Likert Sheet ──────────────────────────────────────────────────────────

    @classmethod
    def _build_likert_sheet(cls, ws, data, cfg, n_resp,
                             title, researcher, subtitle):
        pts    = cfg["points"]
        labels = cfg["labels"]
        values = cfg["values"]

        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "00c9a7"

        # Title block
        ws.merge_cells("A1:K1")
        c = ws["A1"]
        c.value = title or "Tally Report — Part 1 (Likert Scale)"
        c.font  = Font(name="Arial", size=14, bold=True, color="00c9a7")
        c.fill  = PatternFill("solid", fgColor="0d1117")
        c.alignment = cls.CENTER

        ws.merge_cells("A2:K2")
        c = ws["A2"]
        c.value = subtitle or ""
        c.font  = Font(name="Arial", size=10, color="8b949e")
        c.fill  = PatternFill("solid", fgColor="0d1117")
        c.alignment = cls.CENTER

        ws.merge_cells("A3:K3")
        c = ws["A3"]
        c.value = (f"Researcher: {researcher}    |    "
                   f"Date: {datetime.now().strftime('%B %d, %Y')}    |    "
                   f"N = {n_resp} respondents")
        c.font  = Font(name="Arial", size=9, color="8b949e", italic=True)
        c.fill  = PatternFill("solid", fgColor="0d1117")
        c.alignment = cls.CENTER

        ws.row_dimensions[4].height = 6

        # ── Column headers ────────────────────────────────────────────────────
        # Row 5: main headers
        headers = ["#", "Item / Statement"] + labels + ["Total", "WM", "Interpretation"]
        col_widths = [5, 36] + [11]*pts + [8, 8, 18]

        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=5, column=i, value=h)
            c.font      = cls.BOLD_W
            c.fill      = cls.HEADER_FILL
            c.alignment = cls.CENTER
            c.border    = cls.BORDER
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[5].height = 36

        # Sub-header: f (frequency) labels
        freq_start = 3   # column index where scale points start
        for i in range(pts):
            c = ws.cell(row=6, column=freq_start + i, value=f"f{values[i]}")
            c.font      = Font(name="Arial", size=9, color="8b949e", bold=True)
            c.fill      = cls.HEADER_FILL
            c.alignment = cls.CENTER
            c.border    = cls.BORDER
        # Merge the label/n/total cells across row 6 to keep clean
        for col in [1, 2] + list(range(freq_start+pts, freq_start+pts+3)):
            c = ws.cell(row=6, column=col, value="")
            c.fill   = cls.HEADER_FILL
            c.border = cls.BORDER
        ws.row_dimensions[6].height = 16

        # ── Data rows ─────────────────────────────────────────────────────────
        row = 7
        total_wm_sum = 0
        items_with_data = 0

        for idx, item in enumerate(data):
            fill = cls.SUBHEAD_FILL if idx % 2 == 0 else cls.ALT_FILL

            # # column
            cls._cell(ws, row, 1, idx+1, font=Font(name="Arial",size=10,
                      color="8b949e"), fill=fill)
            # Item name
            cls._cell(ws, row, 2, item["name"],
                      font=Font(name="Arial",size=10,color="e6edf3"),
                      fill=fill, align=cls.LEFT)

            freqs  = item["counts"]   # list aligned to values[]
            total  = sum(freqs)
            wm_col = freq_start + pts        # Total column index
            wm_idx = freq_start + pts + 1    # WM column index
            int_col= freq_start + pts + 2    # Interpretation column index

            for i, (f, v) in enumerate(zip(freqs, values)):
                c = ws.cell(row=row, column=freq_start+i, value=f)
                c.font      = Font(name="Arial",size=11,
                                   color="e6edf3" if f > 0 else "30363d",
                                   bold=f>0)
                c.fill      = fill
                c.alignment = cls.CENTER
                c.border    = cls.BORDER

            # Total
            total_formula = (f"=SUM({get_column_letter(freq_start)}{row}:"
                             f"{get_column_letter(freq_start+pts-1)}{row})")
            c = ws.cell(row=row, column=wm_col, value=total_formula)
            c.font=cls.ACCENT_F; c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER

            # Weighted Mean formula  WM = SUMPRODUCT(freqs * values) / total
            val_cells  = ",".join([str(v) for v in values])
            freq_range = ",".join([f"{get_column_letter(freq_start+i)}{row}"
                                    for i in range(pts)])
            # Build WM as Excel SUMPRODUCT
            wm_parts = "+".join(
                [f"{get_column_letter(freq_start+i)}{row}*{values[i]}"
                 for i in range(pts)]
            )
            wm_formula = f"=IFERROR(({wm_parts})/{get_column_letter(wm_col)}{row},0)"
            c = ws.cell(row=row, column=wm_idx, value=wm_formula)
            c.font=cls.ACCENT_F; c.fill=fill; c.alignment=cls.CENTER
            c.border=cls.BORDER
            c.number_format = "0.00"

            # Interpretation
            interp = get_interpretation(item["wm"], cfg["key"]) if item["wm"] else "—"
            c = ws.cell(row=row, column=int_col, value=interp)
            c.font      = Font(name="Arial", size=10,
                               color=("22c55e" if "Agree" in interp and "Dis" not in interp
                                      else "ef4444" if "Disagree" in interp
                                      else "f59e0b"), bold=True)
            c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER

            ws.row_dimensions[row].height = 22
            if item["wm"] and item["wm"] > 0:
                total_wm_sum += item["wm"]
                items_with_data += 1
            row += 1

        # ── Grand Mean row ────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value="OVERALL WEIGHTED MEAN")
        c.font=Font(name="Arial",size=10,bold=True,color="00c9a7")
        c.fill=cls.GRAND_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER

        for i in range(pts):
            c = ws.cell(row=row, column=freq_start+i, value="")
            c.fill=cls.GRAND_FILL; c.border=cls.BORDER

        wm_col_letter = get_column_letter(wm_col+1)   # WM col letter
        # Overall WM = average of all item WMs
        if items_with_data > 0:
            wm_start_row = 7
            wm_end_row   = row - 1
            grand_formula = (f"=IFERROR(AVERAGE({wm_col_letter}{wm_start_row}:"
                             f"{wm_col_letter}{wm_end_row}),0)")
            c = ws.cell(row=row, column=wm_idx, value=grand_formula)
            c.font=Font(name="Arial",size=11,bold=True,color="00c9a7")
            c.fill=cls.GRAND_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER
            c.number_format="0.00"

        grand_interp = get_interpretation(
            total_wm_sum/items_with_data if items_with_data else 0,
            cfg["key"])
        c = ws.cell(row=row, column=int_col, value=grand_interp)
        c.font=Font(name="Arial",size=10,bold=True,color="00c9a7")
        c.fill=cls.GRAND_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER

        c = ws.cell(row=row, column=wm_col, value="")
        c.fill=cls.GRAND_FILL; c.border=cls.BORDER
        ws.row_dimensions[row].height = 24

        # ── Legend ────────────────────────────────────────────────────────────
        row += 2
        ws.merge_cells(f"A{row}:K{row}")
        c = ws.cell(row=row, column=1,
                    value=f"Scale: {cfg['key']}    |    "
                          f"WM = Weighted Mean = Σ(f × v) / N    |    "
                          f"Interpretation based on scale range intervals")
        c.font=Font(name="Arial",size=8,color="8b949e",italic=True)
        c.fill=PatternFill("solid",fgColor="0d1117")
        c.alignment=cls.LEFT

        # Freeze panes
        ws.freeze_panes = f"C7"

    # ── Test Items Sheet ──────────────────────────────────────────────────────

    @classmethod
    def _build_test_sheet(cls, ws, data, n_resp, title, researcher, subtitle):
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "4e9eff"

        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = title or "Tally Report — Part 11 (Test Items)"
        c.font  = Font(name="Arial",size=14,bold=True,color="4e9eff")
        c.fill  = PatternFill("solid",fgColor="0d1117")
        c.alignment = cls.CENTER

        ws.merge_cells("A2:H2")
        c = ws["A2"]
        c.value = subtitle or ""
        c.font  = Font(name="Arial",size=10,color="8b949e")
        c.fill  = PatternFill("solid",fgColor="0d1117")
        c.alignment = cls.CENTER

        ws.merge_cells("A3:H3")
        c = ws["A3"]
        c.value = (f"Researcher: {researcher}    |    "
                   f"Date: {datetime.now().strftime('%B %d, %Y')}    |    "
                   f"N = {n_resp} respondents")
        c.font  = Font(name="Arial",size=9,color="8b949e",italic=True)
        c.fill  = PatternFill("solid",fgColor="0d1117")
        c.alignment = cls.CENTER

        ws.row_dimensions[4].height = 6

        # Headers
        hdrs = ["#","Item","Correct","Wrong","Total",
                "% Correct","% Wrong","Difficulty Index"]
        widths = [5, 32, 12, 12, 10, 12, 12, 18]
        for i, (h, w) in enumerate(zip(hdrs, widths), 1):
            c = ws.cell(row=5, column=i, value=h)
            c.font=cls.BOLD_W; c.fill=cls.HEADER_FILL
            c.alignment=cls.CENTER; c.border=cls.BORDER
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[5].height = 30

        row = 6
        total_correct_all = 0
        total_wrong_all   = 0

        for idx, item in enumerate(data):
            fill = cls.SUBHEAD_FILL if idx % 2 == 0 else cls.ALT_FILL
            correct = item["correct"]
            wrong   = item["wrong"]
            total   = correct + wrong

            total_correct_all += correct
            total_wrong_all   += wrong

            cls._cell(ws, row, 1, idx+1,
                      font=Font(name="Arial",size=10,color="8b949e"),fill=fill)
            cls._cell(ws, row, 2, item["name"],
                      font=Font(name="Arial",size=10,color="e6edf3"),
                      fill=fill, align=cls.LEFT)

            # Correct
            c = ws.cell(row=row, column=3, value=correct)
            c.font=cls.SUCCESS_F; c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER

            # Wrong
            c = ws.cell(row=row, column=4, value=wrong)
            c.font=cls.DANGER_F; c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER

            # Total formula
            c = ws.cell(row=row, column=5, value=f"=C{row}+D{row}")
            c.font=cls.ACCENT_F; c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER

            # % Correct
            c = ws.cell(row=row, column=6,
                        value=f"=IFERROR(C{row}/E{row},0)")
            c.font=cls.SUCCESS_F; c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER
            c.number_format="0.00%"

            # % Wrong
            c = ws.cell(row=row, column=7,
                        value=f"=IFERROR(D{row}/E{row},0)")
            c.font=cls.DANGER_F; c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER
            c.number_format="0.00%"

            # Difficulty Index  p = correct / N
            p_val = correct / n_resp if n_resp > 0 else 0
            diff_label = ("Easy"     if p_val >= 0.76 else
                          "Average"  if p_val >= 0.26 else "Difficult")
            di_formula = f"=IFERROR(C{row}/{n_resp},0)"
            c = ws.cell(row=row, column=8, value=di_formula)
            fcolor = ("22c55e" if diff_label=="Easy" else
                      "f59e0b" if diff_label=="Average" else "ef4444")
            c.font=Font(name="Arial",size=10,bold=True,color=fcolor)
            c.fill=fill; c.alignment=cls.CENTER; c.border=cls.BORDER
            c.number_format="0.00"

            ws.row_dimensions[row].height = 20
            row += 1

        # Totals row
        ws.cell(row=row,column=1,value="").fill=cls.TOTAL_FILL
        c = ws.cell(row=row, column=2, value="TOTALS")
        c.font=cls.BOLD_W; c.fill=cls.TOTAL_FILL; c.alignment=cls.LEFT; c.border=cls.BORDER

        c = ws.cell(row=row, column=3, value=f"=SUM(C6:C{row-1})")
        c.font=cls.SUCCESS_F; c.fill=cls.TOTAL_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER

        c = ws.cell(row=row, column=4, value=f"=SUM(D6:D{row-1})")
        c.font=cls.DANGER_F; c.fill=cls.TOTAL_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER

        c = ws.cell(row=row, column=5, value=f"=SUM(E6:E{row-1})")
        c.font=cls.ACCENT_F; c.fill=cls.TOTAL_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER

        c = ws.cell(row=row, column=6,
                    value=f"=IFERROR(C{row}/E{row},0)")
        c.font=cls.SUCCESS_F; c.fill=cls.TOTAL_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER
        c.number_format="0.00%"

        c = ws.cell(row=row, column=7,
                    value=f"=IFERROR(D{row}/E{row},0)")
        c.font=cls.DANGER_F; c.fill=cls.TOTAL_FILL; c.alignment=cls.CENTER; c.border=cls.BORDER
        c.number_format="0.00%"

        c = ws.cell(row=row, column=8, value="")
        c.fill=cls.TOTAL_FILL; c.border=cls.BORDER
        ws.row_dimensions[row].height = 24

        # Legend
        row += 2
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1,
                    value="Difficulty Index (p): Easy ≥ 0.76 | 0.26–0.75 Average | ≤ 0.25 Difficult")
        c.font=Font(name="Arial",size=8,color="8b949e",italic=True)
        c.fill=PatternFill("solid",fgColor="0d1117"); c.alignment=cls.LEFT

        ws.freeze_panes = "C6"


# ─── Item Row Widget (Likert) ─────────────────────────────────────────────────

class LikertItemRow(ctk.CTkFrame):
    """
    One row in the Likert tally grid.
    Shows: item name | tally buttons per scale | frequency counts | WM | Interpretation
    """

    def __init__(self, master, item_num, item_name, scale_cfg,
                 on_update, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.item_num   = item_num
        self.item_name  = item_name
        self.scale_cfg  = scale_cfg
        self.on_update  = on_update
        self.counts     = [0] * scale_cfg["points"]   # index 0 = highest value
        self.history    = []   # undo stack: list of (index, delta)

        self._build()

    def _build(self):
        pts    = self.scale_cfg["points"]
        colors = self.scale_cfg["colors"]
        labels = self.scale_cfg["labels"]
        values = self.scale_cfg["values"]

        # Item number
        ctk.CTkLabel(self, text=str(self.item_num),
                     font=("Segoe UI",11,"bold"),
                     text_color=TEXT_SEC,
                     width=32).pack(side="left", padx=(4,2))

        # Item name (truncated)
        name_lbl = ctk.CTkLabel(self,
                                 text=self.item_name[:38],
                                 font=("Segoe UI",11),
                                 text_color=TEXT_PRI,
                                 width=240, anchor="w")
        name_lbl.pack(side="left", padx=(0,8))

        # Tally buttons + count display per scale point
        self.count_labels = []
        self.count_vars   = []

        for i in range(pts):
            col = colors[i]
            val = values[i]
            lbl = labels[i].split(" - ")[0] if " - " in labels[i] else str(val)

            # Button
            btn = ctk.CTkButton(self, text=f"+{val}",
                                 fg_color=col, hover_color=col,
                                 text_color="#0d1117",
                                 font=("Segoe UI",11,"bold"),
                                 width=52, height=30, corner_radius=6,
                                 command=lambda idx=i: self._tally(idx))
            btn.pack(side="left", padx=3)

            # Count display
            var = tk.StringVar(value="0")
            self.count_vars.append(var)
            cl = ctk.CTkLabel(self, textvariable=var,
                               font=("Consolas",12,"bold"),
                               text_color=col,
                               width=36)
            cl.pack(side="left", padx=(0,6))
            self.count_labels.append(cl)

        # Divider
        ctk.CTkFrame(self, width=1, height=26,
                     fg_color=BORDER).pack(side="left", padx=4)

        # Total
        self.total_var = tk.StringVar(value="0")
        ctk.CTkLabel(self, text="N:", font=FONT_SML,
                     text_color=TEXT_SEC, width=20).pack(side="left")
        ctk.CTkLabel(self, textvariable=self.total_var,
                     font=("Consolas",12,"bold"),
                     text_color=TEXT_PRI, width=36).pack(side="left", padx=(0,6))

        # WM
        self.wm_var = tk.StringVar(value="—")
        ctk.CTkLabel(self, text="WM:", font=FONT_SML,
                     text_color=TEXT_SEC, width=32).pack(side="left")
        ctk.CTkLabel(self, textvariable=self.wm_var,
                     font=("Consolas",12,"bold"),
                     text_color=ACCENT, width=48).pack(side="left", padx=(0,6))

        # Interpretation
        self.interp_var = tk.StringVar(value="—")
        ctk.CTkLabel(self, textvariable=self.interp_var,
                     font=("Segoe UI",10,"bold"),
                     text_color=WARN, width=130).pack(side="left", padx=4)

        # Undo button
        ctk.CTkButton(self, text="↩",
                      fg_color=BG_PANEL, hover_color=BORDER,
                      text_color=TEXT_SEC,
                      font=("Segoe UI",13), width=32, height=28,
                      corner_radius=6,
                      command=self._undo).pack(side="left", padx=4)

        # Reset item
        ctk.CTkButton(self, text="✕",
                      fg_color="transparent", hover_color=DANGER,
                      text_color=DANGER,
                      font=("Segoe UI",12), width=28, height=28,
                      corner_radius=6,
                      command=self._reset_item).pack(side="left", padx=2)

    def _tally(self, idx):
        self.counts[idx] += 1
        self.history.append(idx)
        self._refresh()

    def _undo(self):
        if not self.history:
            return
        idx = self.history.pop()
        if self.counts[idx] > 0:
            self.counts[idx] -= 1
        self._refresh()

    def _reset_item(self):
        self.counts  = [0] * self.scale_cfg["points"]
        self.history = []
        self._refresh()

    def _refresh(self):
        values = self.scale_cfg["values"]
        total  = sum(self.counts)
        for i, (cnt, var) in enumerate(zip(self.counts, self.count_vars)):
            var.set(str(cnt))
        self.total_var.set(str(total))
        if total > 0:
            wm = sum(c*v for c,v in zip(self.counts, values)) / total
            self.wm_var.set(f"{wm:.2f}")
            self.interp_var.set(get_interpretation(wm, self.scale_cfg["key"]))
        else:
            self.wm_var.set("—")
            self.interp_var.set("—")
        self.on_update()

    def get_data(self):
        values = self.scale_cfg["values"]
        total  = sum(self.counts)
        wm = (sum(c*v for c,v in zip(self.counts,values)) / total
              if total > 0 else 0)
        return {
            "name":   self.item_name,
            "counts": list(self.counts),
            "total":  total,
            "wm":     wm,
        }

    def load_data(self, d):
        self.counts  = list(d.get("counts", [0]*self.scale_cfg["points"]))
        self.history = []
        self._refresh()


# ─── Item Row Widget (Test) ───────────────────────────────────────────────────

class TestItemRow(ctk.CTkFrame):
    """
    One row in the Test Items tally grid.
    Shows: item name | ✓ CORRECT button | ✗ WRONG button | counts | %
    """

    def __init__(self, master, item_num, item_name, n_respondents_var,
                 on_update, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.item_num        = item_num
        self.item_name       = item_name
        self.n_respondents_var = n_respondents_var
        self.on_update       = on_update
        self.correct         = 0
        self.wrong           = 0
        self.history         = []   # 'c' or 'w'

        self._build()

    def _build(self):
        # Number
        ctk.CTkLabel(self, text=str(self.item_num),
                     font=("Segoe UI",11,"bold"),
                     text_color=TEXT_SEC, width=32).pack(side="left", padx=(4,2))

        # Name
        ctk.CTkLabel(self, text=self.item_name[:38],
                     font=("Segoe UI",11), text_color=TEXT_PRI,
                     width=240, anchor="w").pack(side="left", padx=(0,8))

        # CORRECT button
        ctk.CTkButton(self, text="✓  CORRECT",
                      fg_color=SUCCESS, hover_color="#16a34a",
                      text_color="#0d1117",
                      font=("Segoe UI",12,"bold"),
                      width=120, height=30, corner_radius=6,
                      command=self._tally_correct).pack(side="left", padx=4)

        self.correct_var = tk.StringVar(value="0")
        ctk.CTkLabel(self, textvariable=self.correct_var,
                     font=("Consolas",12,"bold"),
                     text_color=SUCCESS, width=36).pack(side="left", padx=(0,8))

        # WRONG button
        ctk.CTkButton(self, text="✗  WRONG",
                      fg_color=DANGER, hover_color="#b91c1c",
                      text_color=TEXT_PRI,
                      font=("Segoe UI",12,"bold"),
                      width=110, height=30, corner_radius=6,
                      command=self._tally_wrong).pack(side="left", padx=4)

        self.wrong_var = tk.StringVar(value="0")
        ctk.CTkLabel(self, textvariable=self.wrong_var,
                     font=("Consolas",12,"bold"),
                     text_color=DANGER, width=36).pack(side="left", padx=(0,8))

        # Divider
        ctk.CTkFrame(self, width=1, height=26,
                     fg_color=BORDER).pack(side="left", padx=4)

        # Total N
        self.total_var = tk.StringVar(value="0")
        ctk.CTkLabel(self, text="N:", font=FONT_SML,
                     text_color=TEXT_SEC, width=20).pack(side="left")
        ctk.CTkLabel(self, textvariable=self.total_var,
                     font=("Consolas",12,"bold"),
                     text_color=TEXT_PRI, width=36).pack(side="left", padx=(0,6))

        # % Correct
        self.pct_var = tk.StringVar(value="—")
        ctk.CTkLabel(self, text="%:", font=FONT_SML,
                     text_color=TEXT_SEC, width=20).pack(side="left")
        ctk.CTkLabel(self, textvariable=self.pct_var,
                     font=("Consolas",12,"bold"),
                     text_color=ACCENT, width=60).pack(side="left", padx=(0,6))

        # Difficulty
        self.diff_var = tk.StringVar(value="—")
        ctk.CTkLabel(self, textvariable=self.diff_var,
                     font=("Segoe UI",10,"bold"),
                     text_color=WARN, width=90).pack(side="left", padx=4)

        # Undo
        ctk.CTkButton(self, text="↩",
                      fg_color=BG_PANEL, hover_color=BORDER,
                      text_color=TEXT_SEC,
                      font=("Segoe UI",13), width=32, height=28,
                      corner_radius=6,
                      command=self._undo).pack(side="left", padx=4)

        # Reset
        ctk.CTkButton(self, text="✕",
                      fg_color="transparent", hover_color=DANGER,
                      text_color=DANGER,
                      font=("Segoe UI",12), width=28, height=28,
                      corner_radius=6,
                      command=self._reset_item).pack(side="left", padx=2)

    def _tally_correct(self):
        self.correct += 1
        self.history.append("c")
        self._refresh()

    def _tally_wrong(self):
        self.wrong += 1
        self.history.append("w")
        self._refresh()

    def _undo(self):
        if not self.history: return
        last = self.history.pop()
        if last == "c" and self.correct > 0:
            self.correct -= 1
        elif last == "w" and self.wrong > 0:
            self.wrong -= 1
        self._refresh()

    def _reset_item(self):
        self.correct = 0; self.wrong = 0; self.history = []
        self._refresh()

    def _refresh(self):
        total = self.correct + self.wrong
        self.correct_var.set(str(self.correct))
        self.wrong_var.set(str(self.wrong))
        self.total_var.set(str(total))
        try:
            n = int(self.n_respondents_var.get())
        except Exception:
            n = total
        if total > 0:
            pct = self.correct / total * 100
            self.pct_var.set(f"{pct:.1f}%")
        else:
            self.pct_var.set("—")
        if n > 0 and self.correct > 0:
            p = self.correct / n
            d = "Easy" if p>=0.76 else "Average" if p>=0.26 else "Difficult"
            dcolor = SUCCESS if d=="Easy" else WARN if d=="Average" else DANGER
            self.diff_var.set(f"p={p:.2f} {d}")
            self.diff_var.set(f"p={p:.2f} {d}")
        else:
            self.diff_var.set("—")
        self.on_update()

    def get_data(self):
        return {
            "name":    self.item_name,
            "correct": self.correct,
            "wrong":   self.wrong,
        }

    def load_data(self, d):
        self.correct = d.get("correct", 0)
        self.wrong   = d.get("wrong",   0)
        self.history = []
        self._refresh()


# ─── Sidebar ──────────────────────────────────────────────────────────────────

class Sidebar(ctk.CTkFrame):
    def __init__(self, master, **kw):
        super().__init__(master, width=260, fg_color=BG_CARD,
                         corner_radius=0, **kw)
        self.pack_propagate(False)
        self._build()

    def _build(self):
        # Logo
        logo = ctk.CTkFrame(self, fg_color=ACCENT, corner_radius=0, height=64)
        logo.pack(fill="x"); logo.pack_propagate(False)
        ctk.CTkLabel(logo, text="  ✎  ", font=("Segoe UI",28,"bold"),
                     text_color="#0d1117", fg_color=ACCENT).pack(expand=True)

        # Scrollable body
        self._scroll = ctk.CTkScrollableFrame(self, fg_color=BG_CARD,
                                               scrollbar_button_color=BORDER,
                                               corner_radius=0)
        self._scroll.pack(fill="both", expand=True)
        s = self._scroll

        ctk.CTkLabel(s, text="Tally Module",
                     font=("Segoe UI",14,"bold"),
                     text_color=TEXT_PRI, fg_color=BG_CARD).pack(pady=(10,2))
        ctk.CTkLabel(s, text="Likert  +  Test Items",
                     font=FONT_TINY, text_color=TEXT_SEC,
                     fg_color=BG_CARD).pack(pady=(0,8))
        divider(s)

        # ── Study info ────────────────────────────────────────────────────────
        sec_label(s, "STUDY TITLE")
        self.title_entry = styled_entry(s, placeholder="e.g. Pilot Test")
        self.title_entry.pack(fill="x", padx=14, pady=(0,4))

        sec_label(s, "RESEARCHER")
        self.researcher_entry = styled_entry(s, placeholder="e.g. Juan Dela Cruz")
        self.researcher_entry.pack(fill="x", padx=14, pady=(0,4))

        sec_label(s, "SUBTITLE / NOTES")
        self.subtitle_entry = styled_entry(s, placeholder="Optional")
        self.subtitle_entry.pack(fill="x", padx=14, pady=(0,4))

        divider(s)

        # ── Part 1 Config ─────────────────────────────────────────────────────
        sec_label(s, "PART 1 — LIKERT CONFIG")

        ctk.CTkLabel(s, text="Scale Type", font=FONT_SML,
                     text_color=TEXT_SEC).pack(anchor="w", padx=18, pady=(4,2))
        self.scale_menu = ctk.CTkOptionMenu(
            s, values=["4-point","5-point","7-point"],
            fg_color=BG_INPUT, button_color=ACCENT,
            button_hover_color="#009e82", text_color=TEXT_PRI,
            dropdown_fg_color=BG_PANEL, dropdown_text_color=TEXT_PRI,
            font=FONT_BODY, height=32, corner_radius=6,
            width=232, dynamic_resizing=False)
        self.scale_menu.set("4-point")
        self.scale_menu.pack(fill="x", padx=14, pady=(0,4))

        ctk.CTkLabel(s, text="Number of Items", font=FONT_SML,
                     text_color=TEXT_SEC).pack(anchor="w", padx=18, pady=(4,2))
        self.likert_n_entry = styled_entry(s, placeholder="e.g. 5", height=32)
        self.likert_n_entry.pack(fill="x", padx=14, pady=(0,4))

        self.build_likert_btn = sidebar_btn(
            s, "⚙  Build Likert Grid",
            fg=ACCENT, hover="#009e82",
            text_color="#0d1117", height=36)
        self.build_likert_btn.pack(fill="x", padx=14, pady=(4,4))

        divider(s)

        # ── Part 11 Config ────────────────────────────────────────────────────
        sec_label(s, "PART 11 — TEST ITEMS CONFIG")

        ctk.CTkLabel(s, text="Number of Items", font=FONT_SML,
                     text_color=TEXT_SEC).pack(anchor="w", padx=18, pady=(4,2))
        self.test_n_entry = styled_entry(s, placeholder="e.g. 15", height=32)
        self.test_n_entry.pack(fill="x", padx=14, pady=(0,4))

        ctk.CTkLabel(s, text="Total Respondents (N)", font=FONT_SML,
                     text_color=TEXT_SEC).pack(anchor="w", padx=18, pady=(4,2))
        self.n_resp_entry = styled_entry(s, placeholder="e.g. 37", height=32)
        self.n_resp_entry.pack(fill="x", padx=14, pady=(0,4))

        self.build_test_btn = sidebar_btn(
            s, "⚙  Build Test Grid",
            fg=ACCENT2, hover="#3b7ddd",
            text_color="#0d1117", height=36)
        self.build_test_btn.pack(fill="x", padx=14, pady=(4,4))

        divider(s)

        pad = {"fill":"x","padx":14,"pady":4}

        self.export_btn = sidebar_btn(
            s, "💾  Export to Excel (.xlsx)",
            fg=SUCCESS, hover="#16a34a",
            text_color="#0d1117", height=40)
        self.export_btn.pack(**pad)

        self.save_session_btn = sidebar_btn(
            s, "📌  Save Session",
            fg=PURPLE, hover="#7c3aed", height=36)
        self.save_session_btn.pack(**pad)

        self.load_session_btn = sidebar_btn(
            s, "📂  Load Session",
            fg="#6d28d9", hover="#5b21b6", height=36)
        self.load_session_btn.pack(**pad)

        self.reset_all_btn = sidebar_btn(
            s, "🗑  Reset All Counts",
            fg=DANGER, hover="#b91c1c", height=36)
        self.reset_all_btn.pack(**pad)

        divider(s)

        self.theme_btn = sidebar_btn(
            s, "☀️  Light Mode",
            fg="#374151", hover="#4b5563",
            font=FONT_BODY, height=32)
        self.theme_btn.pack(fill="x", padx=14, pady=8)

        divider(s)

        # Live stats
        self.stat_frame = ctk.CTkFrame(s, fg_color=BG_PANEL,
                                        corner_radius=8, border_width=1,
                                        border_color=BORDER)
        self.stat_frame.pack(fill="x", padx=14, pady=8)
        ctk.CTkLabel(self.stat_frame, text="LIVE STATS",
                     font=("Segoe UI",10,"bold"),
                     text_color=TEXT_SEC).pack(anchor="w", padx=10, pady=(8,4))

        self.stat_grand_mean = ctk.CTkLabel(
            self.stat_frame, text="Grand Mean: —",
            font=("Consolas",12,"bold"), text_color=ACCENT)
        self.stat_grand_mean.pack(anchor="w", padx=10, pady=2)

        self.stat_total_resp_likert = ctk.CTkLabel(
            self.stat_frame, text="Likert N: 0",
            font=("Consolas",11), text_color=TEXT_SEC)
        self.stat_total_resp_likert.pack(anchor="w", padx=10, pady=2)

        self.stat_test_correct = ctk.CTkLabel(
            self.stat_frame, text="Test Correct: 0",
            font=("Consolas",11), text_color=SUCCESS)
        self.stat_test_correct.pack(anchor="w", padx=10, pady=2)

        self.stat_test_wrong = ctk.CTkLabel(
            self.stat_frame, text="Test Wrong: 0",
            font=("Consolas",11), text_color=DANGER)
        self.stat_test_wrong.pack(anchor="w", padx=10, pady=(2,10))


# ─── Main App ─────────────────────────────────────────────────────────────────

class TallyApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Tally Module — Likert + Test Items")
        self.geometry("1440x860")
        self.minsize(1200,700)
        self.configure(fg_color=BG_DEEP)

        self.likert_rows = []
        self.test_rows   = []
        self.dark_mode   = True

        self._build_ui()

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.sidebar = Sidebar(self)
        self.sidebar.pack(side="left", fill="y")

        # Wire buttons
        self.sidebar.build_likert_btn.configure(command=self.build_likert_grid)
        self.sidebar.build_test_btn.configure(command=self.build_test_grid)
        self.sidebar.export_btn.configure(command=self.export_excel)
        self.sidebar.save_session_btn.configure(command=self.save_session)
        self.sidebar.load_session_btn.configure(command=self.load_session)
        self.sidebar.reset_all_btn.configure(command=self.reset_all)
        self.sidebar.theme_btn.configure(command=self.toggle_theme)

        content = ctk.CTkFrame(self, fg_color=BG_DEEP, corner_radius=0)
        content.pack(side="left", fill="both", expand=True)

        # Header
        hdr = ctk.CTkFrame(content, fg_color=BG_CARD,
                           corner_radius=0, height=64)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="Tally Module",
                     font=FONT_HEAD, text_color=TEXT_PRI).pack(side="left",padx=24)
        self.hdr_sub = ctk.CTkLabel(hdr,
                                     text="Click to tally  •  Auto-computes WM, Total, Grand Mean",
                                     font=("Segoe UI",11), text_color=TEXT_SEC)
        self.hdr_sub.pack(side="right", padx=24)

        # Notebook — two tabs
        nb_style = ttk.Style()
        nb_style.theme_use("default")
        nb_style.configure("Tally.TNotebook",     background=BG_DEEP, borderwidth=0)
        nb_style.configure("Tally.TNotebook.Tab", background=BG_PANEL,
                           foreground=TEXT_SEC, padding=[20,8],
                           font=("Segoe UI",12,"bold"))
        nb_style.map("Tally.TNotebook.Tab",
                     background=[("selected",BG_CARD)],
                     foreground=[("selected",TEXT_PRI)])

        self.nb = ttk.Notebook(content, style="Tally.TNotebook")
        self.nb.pack(fill="both", expand=True, padx=14, pady=12)

        # ── TAB 1: Likert ─────────────────────────────────────────────────────
        self.tab_likert = ctk.CTkFrame(self.nb, fg_color=BG_CARD)
        self.nb.add(self.tab_likert,
                    text="  📋  Part 1 — Likert Scale  ")
        self._build_likert_tab()

        # ── TAB 2: Test Items ─────────────────────────────────────────────────
        self.tab_test = ctk.CTkFrame(self.nb, fg_color=BG_CARD)
        self.nb.add(self.tab_test,
                    text="  📝  Part 11 — Test Items  ")
        self._build_test_tab()

        # Status bar
        bar = ctk.CTkFrame(content, fg_color=BG_CARD,
                           height=28, corner_radius=0)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        self.status_lbl = ctk.CTkLabel(bar, text="Ready — build a grid to start tallying",
                                        font=FONT_TINY, text_color=TEXT_SEC)
        self.status_lbl.pack(side="left", padx=12)
        self.file_lbl = ctk.CTkLabel(bar, text="",
                                      font=FONT_TINY, text_color=TEXT_SEC)
        self.file_lbl.pack(side="right", padx=12)

    # ── Likert Tab skeleton ────────────────────────────────────────────────────

    def _build_likert_tab(self):
        # Column header bar
        self.likert_hdr = ctk.CTkFrame(self.tab_likert,
                                        fg_color=BG_PANEL, corner_radius=6,
                                        height=32)
        self.likert_hdr.pack(fill="x", padx=12, pady=(10,4))
        self.likert_hdr.pack_propagate(False)
        ctk.CTkLabel(self.likert_hdr,
                     text="  #   Item                                        "
                          "← Click button to add 1 →                         "
                          "N        WM      Interpretation      ↩ ✕",
                     font=("Segoe UI",10), text_color=TEXT_SEC).pack(
            side="left", padx=8)

        # Scrollable grid
        self.likert_scroll = ctk.CTkScrollableFrame(
            self.tab_likert, fg_color=BG_CARD,
            scrollbar_button_color=BORDER)
        self.likert_scroll.pack(fill="both", expand=True, padx=12, pady=(0,8))

        # Placeholder
        self.likert_placeholder = ctk.CTkLabel(
            self.likert_scroll,
            text="⚙  Configure Likert scale in the sidebar, then click  ⚙ Build Likert Grid",
            font=FONT_BODY, text_color=TEXT_SEC)
        self.likert_placeholder.pack(pady=60)

        # Summary bar
        self.likert_summary = ctk.CTkFrame(
            self.tab_likert, fg_color=BG_PANEL, corner_radius=8,
            height=38)
        self.likert_summary.pack(fill="x", padx=12, pady=(0,8))
        self.likert_summary.pack_propagate(False)
        self.grand_mean_lbl = ctk.CTkLabel(
            self.likert_summary,
            text="Overall Grand Mean: —",
            font=("Segoe UI",13,"bold"), text_color=ACCENT)
        self.grand_mean_lbl.pack(side="left", padx=16)
        self.likert_n_lbl = ctk.CTkLabel(
            self.likert_summary, text="",
            font=FONT_TINY, text_color=TEXT_SEC)
        self.likert_n_lbl.pack(side="right", padx=16)

    # ── Test Tab skeleton ──────────────────────────────────────────────────────

    def _build_test_tab(self):
        self.test_hdr = ctk.CTkFrame(self.tab_test,
                                      fg_color=BG_PANEL, corner_radius=6,
                                      height=32)
        self.test_hdr.pack(fill="x", padx=12, pady=(10,4))
        self.test_hdr.pack_propagate(False)
        ctk.CTkLabel(self.test_hdr,
                     text="  #   Item                                        "
                          "✓ CORRECT count     ✗ WRONG count     N    %Correct    "
                          "Difficulty      ↩ ✕",
                     font=("Segoe UI",10), text_color=TEXT_SEC).pack(
            side="left", padx=8)

        self.test_scroll = ctk.CTkScrollableFrame(
            self.tab_test, fg_color=BG_CARD,
            scrollbar_button_color=BORDER)
        self.test_scroll.pack(fill="both", expand=True, padx=12, pady=(0,8))

        self.test_placeholder = ctk.CTkLabel(
            self.test_scroll,
            text="⚙  Configure test items in the sidebar, then click  ⚙ Build Test Grid",
            font=FONT_BODY, text_color=TEXT_SEC)
        self.test_placeholder.pack(pady=60)

        # Summary bar
        self.test_summary = ctk.CTkFrame(
            self.tab_test, fg_color=BG_PANEL, corner_radius=8, height=38)
        self.test_summary.pack(fill="x", padx=12, pady=(0,8))
        self.test_summary.pack_propagate(False)
        self.test_correct_lbl = ctk.CTkLabel(
            self.test_summary, text="Total Correct: 0",
            font=("Segoe UI",13,"bold"), text_color=SUCCESS)
        self.test_correct_lbl.pack(side="left", padx=16)
        self.test_wrong_lbl = ctk.CTkLabel(
            self.test_summary, text="Total Wrong: 0",
            font=("Segoe UI",13,"bold"), text_color=DANGER)
        self.test_wrong_lbl.pack(side="left", padx=16)
        self.test_pct_lbl = ctk.CTkLabel(
            self.test_summary, text="",
            font=FONT_TINY, text_color=TEXT_SEC)
        self.test_pct_lbl.pack(side="right", padx=16)

    # ── Build grids ───────────────────────────────────────────────────────────

    def build_likert_grid(self):
        try:
            n = int(self.sidebar.likert_n_entry.get())
            if not 1 <= n <= 100:
                messagebox.showwarning("Invalid","Enter 1–100 items."); return
        except ValueError:
            messagebox.showerror("Error","Enter a valid number."); return

        scale_key = self.sidebar.scale_menu.get()
        cfg = dict(SCALE_CONFIGS[scale_key])
        cfg["key"] = scale_key

        # Clear existing
        for w in self.likert_scroll.winfo_children():
            w.destroy()
        self.likert_rows.clear()

        if hasattr(self,"likert_placeholder"):
            try: self.likert_placeholder.destroy()
            except: pass

        for i in range(n):
            row = LikertItemRow(
                self.likert_scroll,
                item_num=i+1,
                item_name=f"Item {i+1}",
                scale_cfg=cfg,
                on_update=self._refresh_likert_summary)
            row.pack(fill="x", pady=3, padx=4)
            self.likert_rows.append(row)

        self._refresh_likert_summary()
        self.status_lbl.configure(
            text=f"✓ Likert grid built: {n} items  ({scale_key})")
        self.nb.select(0)

    def build_test_grid(self):
        try:
            n = int(self.sidebar.test_n_entry.get())
            if not 1 <= n <= 100:
                messagebox.showwarning("Invalid","Enter 1–100 items."); return
        except ValueError:
            messagebox.showerror("Error","Enter a valid number."); return

        for w in self.test_scroll.winfo_children():
            w.destroy()
        self.test_rows.clear()

        if hasattr(self,"test_placeholder"):
            try: self.test_placeholder.destroy()
            except: pass

        for i in range(n):
            row = TestItemRow(
                self.test_scroll,
                item_num=i+1,
                item_name=f"Item {i+1}",
                n_respondents_var=self.sidebar.n_resp_entry,
                on_update=self._refresh_test_summary)
            row.pack(fill="x", pady=3, padx=4)
            self.test_rows.append(row)

        self._refresh_test_summary()
        self.status_lbl.configure(
            text=f"✓ Test grid built: {n} items")
        self.nb.select(1)

    # ── Live summary refreshers ────────────────────────────────────────────────

    def _refresh_likert_summary(self):
        if not self.likert_rows:
            self.grand_mean_lbl.configure(text="Overall Grand Mean: —")
            return
        wms = [r.get_data()["wm"] for r in self.likert_rows
               if r.get_data()["wm"] > 0]
        grand = sum(wms) / len(wms) if wms else 0
        total_n = sum(r.get_data()["total"] for r in self.likert_rows)

        if grand > 0:
            key = self.sidebar.scale_menu.get()
            interp = get_interpretation(grand, key)
            self.grand_mean_lbl.configure(
                text=f"Overall Grand Mean: {grand:.2f}  —  {interp}")
        else:
            self.grand_mean_lbl.configure(text="Overall Grand Mean: —")
        self.likert_n_lbl.configure(
            text=f"Items with data: {len(wms)} / {len(self.likert_rows)}   "
                 f"Total tallied: {total_n}")

        # Sidebar live stats
        self.sidebar.stat_grand_mean.configure(
            text=f"Grand Mean: {grand:.2f}" if grand>0 else "Grand Mean: —")
        self.sidebar.stat_total_resp_likert.configure(
            text=f"Likert tallied: {total_n}")

    def _refresh_test_summary(self):
        if not self.test_rows:
            return
        total_c = sum(r.correct for r in self.test_rows)
        total_w = sum(r.wrong   for r in self.test_rows)
        total   = total_c + total_w
        self.test_correct_lbl.configure(text=f"Total Correct: {total_c}")
        self.test_wrong_lbl.configure(text=f"Total Wrong: {total_w}")
        pct = total_c/total*100 if total>0 else 0
        self.test_pct_lbl.configure(
            text=f"Overall % Correct: {pct:.1f}%   Total tallied: {total}")
        self.sidebar.stat_test_correct.configure(
            text=f"Test Correct: {total_c}")
        self.sidebar.stat_test_wrong.configure(
            text=f"Test Wrong: {total_w}")

    # ── Export ────────────────────────────────────────────────────────────────

    def export_excel(self):
        if not self.likert_rows and not self.test_rows:
            messagebox.showwarning("No Data",
                                   "Build at least one grid first."); return
        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"TallyReport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not fp: return

        try:
            scale_key = self.sidebar.scale_menu.get()
            cfg = dict(SCALE_CONFIGS[scale_key])
            cfg["key"] = scale_key

            likert_data = [r.get_data() for r in self.likert_rows]
            test_data   = [r.get_data() for r in self.test_rows]

            # N respondents for test
            try:
                n_resp = int(self.sidebar.n_resp_entry.get())
            except Exception:
                n_resp = (max((r.correct+r.wrong for r in self.test_rows), default=0)
                          if self.test_rows else 0)

            n_likert = (max((sum(r.counts) for r in self.likert_rows), default=0)
                        if self.likert_rows else 0)

            ExcelExporter.export(
                filename=fp,
                likert_data=likert_data,
                test_data=test_data,
                likert_cfg=cfg,
                n_respondents_likert=n_likert,
                n_respondents_test=n_resp,
                title=self.sidebar.title_entry.get().strip() or "Tally Report",
                researcher=self.sidebar.researcher_entry.get().strip(),
                subtitle=self.sidebar.subtitle_entry.get().strip(),
            )
            self.file_lbl.configure(
                text=f"Saved: {os.path.basename(fp)}")
            self.status_lbl.configure(
                text=f"✓ Excel exported: {os.path.basename(fp)}")
            messagebox.showinfo("Exported",
                f"Excel file saved!\n{fp}\n\n"
                f"Sheets:\n"
                f"  • Part 1 — Likert  ({len(likert_data)} items)\n"
                f"  • Part 11 — Test Items  ({len(test_data)} items)")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ── Session Save / Load ────────────────────────────────────────────────────

    def save_session(self):
        if not self.likert_rows and not self.test_rows:
            messagebox.showwarning("No Data","Nothing to save."); return
        fp = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON Session","*.json")],
            initialfile=f"TallySession_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        if not fp: return
        try:
            session = {
                "meta": {
                    "title":      self.sidebar.title_entry.get(),
                    "researcher": self.sidebar.researcher_entry.get(),
                    "subtitle":   self.sidebar.subtitle_entry.get(),
                    "scale":      self.sidebar.scale_menu.get(),
                    "n_resp":     self.sidebar.n_resp_entry.get(),
                    "saved_at":   datetime.now().isoformat(),
                },
                "likert": [r.get_data() for r in self.likert_rows],
                "test":   [r.get_data() for r in self.test_rows],
            }
            with open(fp,"w") as f:
                json.dump(session, f, indent=2)
            self.status_lbl.configure(
                text=f"✓ Session saved: {os.path.basename(fp)}")
            messagebox.showinfo("Saved",f"Session saved:\n{fp}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def load_session(self):
        fp = filedialog.askopenfilename(
            filetypes=[("JSON Session","*.json"),("All","*.*")],
            title="Load Tally Session")
        if not fp: return
        try:
            with open(fp) as f:
                session = json.load(f)
            meta = session.get("meta",{})

            # Restore meta
            self.sidebar.title_entry.delete(0,"end")
            self.sidebar.title_entry.insert(0, meta.get("title",""))
            self.sidebar.researcher_entry.delete(0,"end")
            self.sidebar.researcher_entry.insert(0, meta.get("researcher",""))
            self.sidebar.subtitle_entry.delete(0,"end")
            self.sidebar.subtitle_entry.insert(0, meta.get("subtitle",""))
            self.sidebar.scale_menu.set(meta.get("scale","4-point"))
            self.sidebar.n_resp_entry.delete(0,"end")
            self.sidebar.n_resp_entry.insert(0, meta.get("n_resp",""))

            # Rebuild + load Likert
            ldata = session.get("likert",[])
            if ldata:
                self.sidebar.likert_n_entry.delete(0,"end")
                self.sidebar.likert_n_entry.insert(0, str(len(ldata)))
                self.build_likert_grid()
                for row, d in zip(self.likert_rows, ldata):
                    row.item_name = d.get("name", row.item_name)
                    row.load_data(d)

            # Rebuild + load Test
            tdata = session.get("test",[])
            if tdata:
                self.sidebar.test_n_entry.delete(0,"end")
                self.sidebar.test_n_entry.insert(0, str(len(tdata)))
                self.build_test_grid()
                for row, d in zip(self.test_rows, tdata):
                    row.item_name = d.get("name", row.item_name)
                    row.load_data(d)

            self.status_lbl.configure(
                text=f"✓ Session loaded: {os.path.basename(fp)}")
            messagebox.showinfo("Loaded",
                f"Session loaded!\n"
                f"Likert items: {len(ldata)}\n"
                f"Test items:   {len(tdata)}")
        except Exception as e:
            messagebox.showerror("Load Error", str(e))

    # ── Reset ──────────────────────────────────────────────────────────────────

    def reset_all(self):
        if not messagebox.askyesno("Reset All",
                                    "Reset ALL tally counts?\n"
                                    "This cannot be undone."):
            return
        for r in self.likert_rows: r._reset_item()
        for r in self.test_rows:   r._reset_item()
        self.status_lbl.configure(text="All counts reset.")

    # ── Theme ──────────────────────────────────────────────────────────────────

    def toggle_theme(self):
        if self.dark_mode:
            ctk.set_appearance_mode("light")
            self.sidebar.theme_btn.configure(text="🌙  Dark Mode")
            self.dark_mode = False
        else:
            ctk.set_appearance_mode("dark")
            self.sidebar.theme_btn.configure(text="☀️  Light Mode")
            self.dark_mode = True


# ─── Entry ────────────────────────────────────────────────────────────────────

def main():
    app = TallyApp()
    app.mainloop()

if __name__ == "__main__":
    main()